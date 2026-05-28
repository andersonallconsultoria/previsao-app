from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
import requests
from decouple import config
from .decorators import token_required
from collections import defaultdict
import logging
import json
from datetime import datetime
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Cache dinâmico para configurações
_config_cache = {}
_config_cache_timestamp = 0

def get_dynamic_config(key, default=None):
    """
    Obtém configuração dinamicamente do arquivo .env
    com cache para performance
    """
    global _config_cache, _config_cache_timestamp
    
    # Verifica se o arquivo .env foi modificado
    try:
        env_path = '.env'
        if os.path.exists(env_path):
            current_timestamp = os.path.getmtime(env_path)
            
            # Se o arquivo foi modificado, limpa o cache
            if current_timestamp > _config_cache_timestamp:
                _config_cache.clear()
                _config_cache_timestamp = current_timestamp
                logger.info("🔄 Cache de configurações limpo - arquivo .env modificado")
        
        # Se não está no cache, lê do arquivo
        if key not in _config_cache:
            if os.path.exists(env_path):
                with open(env_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#') and '=' in line:
                            k, v = line.split('=', 1)
                            _config_cache[k] = v
                            if logger.isEnabledFor(logging.DEBUG):
                                logger.debug(f"📝 Config carregada: {k}={v}")
            
            # Se ainda não encontrou, usa o config padrão
            if key not in _config_cache:
                _config_cache[key] = config(key, default=default)
        
        return _config_cache.get(key, default)
        
    except Exception as e:
        logger.warning(f"⚠️ Erro ao ler configuração dinâmica: {e}")
        return config(key, default=default)

# Filtro de template
from django.template.defaulttags import register
@register.filter
def get_item(dictionary, key):
    return dictionary.get(str(key), {})

# Função auxiliar para logs JSON bonitos
def log_json_pretty(obj, title=None):
    if title:
        logger.info(f"{title}\n{json.dumps(obj, indent=2, ensure_ascii=False)}")
    else:
        logger.info(json.dumps(obj, indent=2, ensure_ascii=False))

# Token OAuth2
def gerar_token(username, password):
    url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-auth/oauth/token"
    payload = {
        'username': username,
        'password': password,
        'grant_type': 'password',
        'client_id': get_dynamic_config('CLIENT_ID'),
        'client_secret': get_dynamic_config('CLIENT_SECRET'),
    }
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    
    try:
        # Adicionar timeout para evitar esperas muito longas
        response = requests.post(url, data=payload, headers=headers, timeout=10)
        return response.json()
    except requests.exceptions.Timeout:
        logger.error(f"⏰ Timeout ao conectar com {url}")
        return {'error': 'timeout', 'error_description': 'Timeout de conexão'}
    except requests.exceptions.ConnectTimeout:
        logger.error(f"⏰ Connect timeout ao conectar com {url}")
        return {'error': 'connect_timeout', 'error_description': 'Timeout de conexão'}
    except requests.exceptions.ConnectionError:
        logger.error(f"🔌 Erro de conexão com {url}")
        return {'error': 'connection_error', 'error_description': 'Erro de conexão'}
    except Exception as e:
        logger.error(f"❌ Erro inesperado ao gerar token: {str(e)}")
        return {'error': 'unknown', 'error_description': f'Erro inesperado: {str(e)}'}

# Login
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        try:
            token_data = gerar_token(username, password)
            if 'access_token' in token_data:
                request.session['token'] = token_data['access_token']
                request.session['username'] = username.upper()  # Armazenar username em maiúsculas
                return redirect('painel')
            else:
                # Erro de autenticação da API
                error_msg = token_data.get('error_description', 'Login inválido')
                if 'invalid_grant' in error_msg.lower():
                    error_msg = 'Usuário ou senha incorretos'
                elif 'timeout' in error_msg.lower() or 'connection' in error_msg.lower():
                    error_msg = 'Erro de conexão com o servidor. Verifique a configuração de conexão.'
                
                return render(request, 'login.html', {
                    'erro': error_msg,
                    'erro_tipo': 'auth'
                })
        except requests.exceptions.ConnectTimeout:
            return render(request, 'login.html', {
                'erro': '❌ Erro de conexão: Timeout ao conectar com o servidor. Verifique se o IP e porta estão corretos.',
                'erro_tipo': 'connection'
            })
        except requests.exceptions.ConnectionError:
            return render(request, 'login.html', {
                'erro': '❌ Erro de conexão: Não foi possível conectar com o servidor. Verifique se o servidor está rodando e acessível.',
                'erro_tipo': 'connection'
            })
        except Exception as e:
            logger.error(f"❌ Erro inesperado no login: {str(e)}")
            return render(request, 'login.html', {
                'erro': f'❌ Erro inesperado: {str(e)}',
                'erro_tipo': 'unknown'
            })
    
    # Pass current API_BASE_URL to template
    context = {
        'api_base_url': get_dynamic_config('API_BASE_URL', default='http://200.141.41.20:8086')
    }
    return render(request, 'login.html', context)

# Painel principal
@token_required
def painel_view(request):
    token = request.session['token']
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }

    username = request.session.get('username', '')
    centros_permitidos_ids = get_centros_permitidos(username)
    sem_centros_permitidos = False
    mensagem_bloqueio = None
    tem_lista_permitida = bool(centros_permitidos_ids)

    ano_atual = datetime.now().year

    def post_api(endpoint):
        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/{endpoint}"
        try:
            resp = requests.post(url, json={"page": 1}, headers=headers)
            if resp.status_code == 200:
                return resp.json().get("data", [])
            else:
                logger.warning(f"⚠️ {endpoint} retornou {resp.status_code}")
        except Exception as e:
            logger.exception(f"❌ Erro ao buscar {endpoint}")
        return []

    empresas = post_api('cadastro_empresa')
    centros = post_api('cadastro_centroresultados')
    centro_preselect_id = None
    # Filtrar centros conforme permissões do usuário
    if centros_permitidos_ids:
        centros = [
            c for c in centros
            if str(c.get('idcentroresultado')) and int(c.get('idcentroresultado')) in centros_permitidos_ids
        ]
        if len(centros) == 1:
            try:
                centro_preselect_id = int(centros[0].get('idcentroresultado'))
            except Exception:
                centro_preselect_id = None
    else:
        # Opção B: bloquear acesso se não houver centros vinculados
        sem_centros_permitidos = True
        mensagem_bloqueio = "Usuário não tem centro de resultados relacionado. Contate o administrador."
        centros = []

    resultados = []
    agrupado = {}
    meses_lista = [(f"{i:02d}", f"{i:02d}") for i in range(1, 13)]

    if request.method == 'POST':
        # Bloqueado: não processa pesquisa
        if sem_centros_permitidos:
            return render(request, 'painel.html', {
                'empresas': empresas,
                'centros': centros,
                'resultados': resultados,
                'agrupado': {},
                'agrupado_por_centro': {},
                'totalizadores_centro': {},
                'totais_anuais_conta': {},
                'agrupar_por_centro': False,
                'meses': meses_lista,
                'ano_atual': ano_atual,
                'pode_gerenciar_usuarios': usuario_pode_gerenciar_usuarios(username),
                'tem_acesso_configuracao': usuario_tem_acesso_configuracao(username),
                'sem_centros_permitidos': sem_centros_permitidos,
                'mensagem_bloqueio': mensagem_bloqueio,
                'tem_lista_permitida': tem_lista_permitida,
            })

        ano = int(request.POST.get("ano") or ano_atual)
        empresa = request.POST.get("empresa") or None
        centro = request.POST.get("centro") or None

        # Validar se o centro solicitado é permitido
        if tem_lista_permitida and not centro:
            mensagem_bloqueio = "Selecione um centro de resultado permitido para continuar."
            return render(request, 'painel.html', {
                'empresas': empresas,
                'centros': centros,
                'resultados': resultados,
                'agrupado': {},
                'agrupado_por_centro': {},
                'totalizadores_centro': {},
                'totais_anuais_conta': {},
                'agrupar_por_centro': False,
                'meses': meses_lista,
                'ano_atual': ano_atual,
                'pode_gerenciar_usuarios': usuario_pode_gerenciar_usuarios(username),
                'tem_acesso_configuracao': usuario_tem_acesso_configuracao(username),
                'sem_centros_permitidos': sem_centros_permitidos,
                'mensagem_bloqueio': mensagem_bloqueio,
                'tem_lista_permitida': tem_lista_permitida,
            })

        if centro and centros_permitidos_ids and int(centro) not in centros_permitidos_ids:
            mensagem_bloqueio = "Você não tem permissão para acessar este centro de resultado."
            return render(request, 'painel.html', {
                'empresas': empresas,
                'centros': centros,
                'resultados': resultados,
                'agrupado': {},
                'agrupado_por_centro': {},
                'totalizadores_centro': {},
                'totais_anuais_conta': {},
                'agrupar_por_centro': False,
                'meses': meses_lista,
                'ano_atual': ano_atual,
                'pode_gerenciar_usuarios': usuario_pode_gerenciar_usuarios(username),
                'tem_acesso_configuracao': usuario_tem_acesso_configuracao(username),
                'sem_centros_permitidos': sem_centros_permitidos,
                'mensagem_bloqueio': mensagem_bloqueio,
                'tem_lista_permitida': tem_lista_permitida,
            })

        logger.info(f"🔎 Filtros aplicados: ano={ano}, empresa={empresa}, centro={centro}")

        payload = {
            "page": 1,
            "limit": 1000,
            "clausulas": [
                {"campo": "anoreferencia", "operadorlogico": "AND", "operador": "IGUAL", "valor": ano}
            ]
        }

        # Sempre adicionar cláusula de empresa
        if empresa and empresa.strip():
            try:
                emp_int = int(empresa)
                payload["clausulas"].append({
                    "campo": "idempfiltro", "operadorlogico": "AND", "operador": "IGUAL", "valor": emp_int
                })
                logger.info(f"🏢 Empresa específica selecionada: {emp_int}")
            except ValueError:
                logger.warning(f"⚠️ Empresa inválida: {empresa}")
                # Se empresa inválida, enviar null para buscar todas
                payload["clausulas"].append({
                    "campo": "idempfiltro", "operadorlogico": "AND", "operador": "IGUAL", "valor": None
                })
                logger.info("🏢 Empresa inválida - enviando null (buscar todas)")
        else:
            # Se empresa estiver vazia ou for "Todas", enviar null com operador IGUAL
            payload["clausulas"].append({
                "campo": "idempfiltro", "operadorlogico": "AND", "operador": "IGUAL", "valor": None
            })
            logger.info("🏢 Filtro de empresa: 'Todas' selecionado - enviando null")

        # Sempre adicionar a cláusula de centro de resultados
        if centro and centro.strip():
            try:
                centro_int = int(centro)
                payload["clausulas"].append({
                    "campo": "idcentroresultadofiltro", "operadorlogico": "AND", "operador": "IGUAL", "valor": centro_int
                })
                # Se centro específico foi selecionado, não agrupar por centro
                agrupar_por_centro = False
                logger.info(f"🎯 Centro específico selecionado: {centro_int}")
            except ValueError:
                logger.warning(f"⚠️ Centro inválido: {centro}")
                # Se centro inválido, enviar null
                payload["clausulas"].append({
                    "campo": "idcentroresultadofiltro", "operadorlogico": "AND", "operador": "IGUAL", "valor": None
                })
                agrupar_por_centro = True
                logger.info("🔄 Centro inválido - enviando null")
        else:
            # Se centro estiver vazio ou for "Todos", enviar null
            payload["clausulas"].append({
                "campo": "idcentroresultadofiltro", "operadorlogico": "AND", "operador": "IGUAL", "valor": None
            })
            logger.info("🏢 Filtro de centro: 'Todos' selecionado - enviando null")
            agrupar_por_centro = True

        logger.info(f"📤 Enviando payload para /centro_resultado_bi:")
        logger.info(f"🔍 Filtros aplicados:")
        logger.info(f"   - Ano: {ano}")
        logger.info(f"   - Empresa: {empresa if empresa else 'Todas'}")
        logger.info(f"   - Centro: {centro if centro and centro.strip() else 'Todos (null)'}")
        logger.info(f"   - Agrupar por centro: {agrupar_por_centro}")
        log_json_pretty(payload)

        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/centro_resultado_bi"
        try:
            # Implementar paginação para buscar todas as páginas
            resultados = []
            page = 1
            total_registros = 0
            
            while True:
                payload["page"] = page
                logger.info(f"📄 Buscando página {page}...")
                
                resp = requests.post(url, json=payload, headers=headers)
                logger.info(f"📥 Status da requisição (página {page}): {resp.status_code}")
                
                if resp.status_code == 200:
                    response_data = resp.json()
                    page_data = response_data.get("data", [])
                    total_registros = response_data.get("total", 0)
                    has_next = response_data.get("hasNext", False)
                    
                    resultados.extend(page_data)
                    logger.info(f"✅ Página {page}: {len(page_data)} registros recebidos")
                    logger.info(f"📊 Total acumulado: {len(resultados)} de {total_registros}")
                    
                    if not has_next:
                        logger.info(f"🏁 Última página alcançada: {page}")
                        break
                    
                    page += 1
                else:
                    logger.warning(f"⚠️ Erro na página {page}: {resp.text}")
                    break
            
            # Amostra para a tela de debug — guardar lista inteira na sessão estoura cookie/DB
            request.session['resultados_debug'] = resultados[:50]
            logger.info(f"✅ Total final de registros: {len(resultados)}")
            log_json_pretty(resultados[:3], "📊 Primeiros resultados:")
            
            # Verificar se "Todas" empresas foi selecionada (empresa vazia ou None)
            agrupar_por_empresa = not (empresa and empresa.strip())
            
            if agrupar_por_empresa:
                logger.info("🏢 'Todas' empresas selecionada - agrupando por empresa")
            
            resultado_agrupado = agrupar_resultados(resultados, agrupar_por_centro, agrupar_por_empresa)
            
            # Salvar dados agrupados na sessão para exportação rápida (sem nova requisição à API)
            request.session['dados_exportacao'] = {
                'resultado_agrupado': resultado_agrupado,
                'agrupar_por_centro': agrupar_por_centro,
                'agrupar_por_empresa': agrupar_por_empresa,
                'ano': ano,
                'empresa': empresa,
                'centro': centro,
            }
            logger.info("💾 Dados salvos na sessão para exportação rápida")
                
            if agrupar_por_centro:
                agrupado_por_centro = resultado_agrupado['agrupado_por_centro']
                totalizadores_centro = resultado_agrupado['totalizadores_centro']
                agrupado = {}  # Não usado quando agrupado por centro
                agrupado_por_empresa = {}  # Não usado quando agrupado por centro
                
                for centro, contas in agrupado_por_centro.items():
                    for conta, meses in contas.items():
                        for mes, valores in meses.items():
                            logger.info(f"📌 Centro: {centro} | Conta: {conta} | Mês: {mes} | Dados: {valores}")
                
                logger.info(f"📦 Total de centros agrupados: {len(agrupado_por_centro)}")
                logger.info(f"📦 Totalizadores por centro: {len(totalizadores_centro)}")
                log_json_pretty(list(agrupado_por_centro.keys())[:5], "🔑 Centros agrupados:")
            elif agrupar_por_empresa:
                agrupado_por_empresa = resultado_agrupado['agrupado_por_empresa']
                totalizadores_centro = resultado_agrupado['totalizadores_centro']
                totais_anuais_conta = resultado_agrupado.get('totais_anuais_conta', {})
                totalizadores_empresa = resultado_agrupado.get('totalizadores_empresa', {})
                totalizadores_geral = resultado_agrupado.get('totalizadores_geral', {})
                agrupado = {}  # Não usado quando agrupado por empresa
                agrupado_por_centro = {}  # Não usado quando agrupado por empresa
                
                logger.info(f"📦 Total de empresas agrupadas: {len(agrupado_por_empresa)}")
                logger.info(f"📦 Lista de empresas encontradas: {list(agrupado_por_empresa.keys())}")
                for empresa, contas in agrupado_por_empresa.items():
                    logger.info(f"🏢 Empresa: '{empresa}' | Total de contas: {len(contas)}")
                    # Log DEBUG apenas se habilitado e para primeiros itens (evitar sobrecarga)
                    if logger.isEnabledFor(logging.DEBUG):
                        count = 0
                        for conta, meses in contas.items():
                            if count >= 3:  # Limitar a 3 contas
                                break
                            for mes, valores in list(meses.items())[:3]:  # Limitar a 3 meses por conta
                                logger.debug(f"📌 Empresa: {empresa} | Conta: {conta} | Mês: {mes} | Dados: {valores}")
                            count += 1
                
                logger.info(f"📦 Totalizadores por centro: {len(totalizadores_centro)}")
                logger.info(f"📦 Totalizadores por empresa: {len(totalizadores_empresa)}")
                log_json_pretty(list(agrupado_por_empresa.keys())[:5], "🔑 Empresas agrupadas:")
            else:
                agrupado = resultado_agrupado['agrupado']
                totalizadores_centro = resultado_agrupado['totalizadores_centro']
                totais_anuais_conta = resultado_agrupado.get('totais_anuais_conta', {})
                agrupado_por_centro = {}  # Não usado quando agrupado por conta
                agrupado_por_empresa = {}  # Não usado quando agrupado por conta
                
                for conta, meses in agrupado.items():
                    for mes, valores in meses.items():
                        logger.info(f"📌 Conta: {conta} | Mês: {mes} | Dados: {valores}")
                
                logger.info(f"📦 Total de contas agrupadas: {len(agrupado)}")
                logger.info(f"📦 Totalizadores por centro: {len(totalizadores_centro)}")
                log_json_pretty(list(agrupado.keys())[:5], "🔑 Contas agrupadas:")
            
            log_json_pretty(totalizadores_centro, "💰 Totalizadores por centro:")
        except Exception:
            logger.exception("❌ Erro na requisição ao endpoint /centro_resultado_bi")
            

    # Verificar permissões do usuário
    pode_gerenciar_usuarios = usuario_pode_gerenciar_usuarios(username)
    tem_acesso_configuracao = usuario_tem_acesso_configuracao(username)
    
    return render(request, 'painel.html', {
        'empresas': empresas,
        'centros': centros,
        'resultados': resultados,
        'agrupado': agrupado if 'agrupado' in locals() else {},
        'agrupado_por_centro': agrupado_por_centro if 'agrupado_por_centro' in locals() else {},
        'agrupado_por_empresa': agrupado_por_empresa if 'agrupado_por_empresa' in locals() else {},
        'totalizadores_centro': totalizadores_centro if 'totalizadores_centro' in locals() else {},
        'totalizadores_empresa': totalizadores_empresa if 'totalizadores_empresa' in locals() else {},
        'totalizadores_geral': totalizadores_geral if 'totalizadores_geral' in locals() else {},
        'totais_anuais_conta': totais_anuais_conta if 'totais_anuais_conta' in locals() else {},
        'agrupar_por_centro': agrupar_por_centro if 'agrupar_por_centro' in locals() else False,
        'agrupar_por_empresa': agrupar_por_empresa if 'agrupar_por_empresa' in locals() else False,
        'meses': meses_lista,
        'ano_atual': ano_atual,
        'pode_gerenciar_usuarios': pode_gerenciar_usuarios,
        'tem_acesso_configuracao': tem_acesso_configuracao,
        'sem_centros_permitidos': sem_centros_permitidos,
        'mensagem_bloqueio': mensagem_bloqueio,
        'tem_lista_permitida': tem_lista_permitida,
        'centro_preselect_id': centro_preselect_id,
    })

# Agrupamento dos dados por conta e mês com totalizadores por centro de resultados
def agrupar_resultados(dados, agrupar_por_centro=False, agrupar_por_empresa=False):
    from collections import defaultdict

    if agrupar_por_centro:
        # Agrupamento por centro de resultados e depois por conta
        agrupado_por_centro = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
        totalizadores_centro = defaultdict(lambda: defaultdict(float))
    elif agrupar_por_empresa:
        # Agrupamento por empresa quando "Todas" empresas foi selecionada
        agrupado_por_empresa_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
        agrupado = {}  # Não usado quando agrupado por empresa
        totalizadores_centro = defaultdict(lambda: defaultdict(float))
        totais_anuais_conta = defaultdict(lambda: defaultdict(float))
        totalizadores_empresa = defaultdict(lambda: defaultdict(float))  # Totalizadores por empresa
    else:
        # Agrupamento tradicional por conta
        agrupado = defaultdict(lambda: defaultdict(dict))
        totalizadores_centro = defaultdict(lambda: defaultdict(float))
        totais_anuais_conta = defaultdict(lambda: defaultdict(float))

    for item in dados:
        conta = item.get("contabil")
        mes_raw = item.get("mesNum")
        # Tentar diferentes campos possíveis para centro de resultados
        centro_resultado = (
            item.get("centroresultados") or 
            item.get("centroResultados") or 
            item.get("centro_resultados") or 
            item.get("centro") or 
            item.get("centroResultado") or 
            item.get("descrcentroresultado") or  # Adicionar este campo
            item.get("CENTRO_RESULTADOS") or  # Campo da função DB2
            "Sem Centro"
        )
        # Obter informação da empresa - tentar diferentes campos possíveis
        empresa_info = (
            item.get("empresa") or 
            item.get("EMPRESA") or 
            item.get("empresaNome") or
            item.get("EMPRESA_NOME") or
            (f"{item.get('idempresa', item.get('IDEMPRESA', ''))} - {item.get('empresaNome', item.get('EMPRESA_NOME', 'Empresa'))}" if item.get('idempresa') or item.get('IDEMPRESA') else "Empresa não identificada")
        )
        # Log para debug (apenas se DEBUG estiver habilitado)
        if logger.isEnabledFor(logging.DEBUG) and agrupar_por_empresa and len(agrupado_por_empresa_dict) <= 3:
            logger.debug(f"🏢 Campo empresa encontrado: {empresa_info} | idempresa: {item.get('idempresa')} | empresa: {item.get('empresa')}")

        # Validação dos dados obrigatórios
        if not conta:
            logger.warning(f"⚠️ Conta contábil ausente no item: {item}")
            continue

        if mes_raw is None:
            logger.warning(f"⚠️ Mês de referência ausente no item: {item}")
            continue

        try:
            mes = str(int(mes_raw)).zfill(2)  # Garantir que seja string "01", "02", ..., "12"
        except Exception:
            logger.exception(f"❌ Erro ao processar o mês: {mes_raw}")
            continue

        try:
            ano_anterior = item.get("anoAnterior", 0)
            previsto = item.get("valorPrevisto", 0)
            realizado = item.get("valorRealizado", 0)

            if agrupar_por_centro:
                # Agrupar por centro de resultados primeiro
                agrupado_por_centro[centro_resultado][conta][mes] = {
                    'ano_anterior': ano_anterior,
                    'previsto': previsto,
                    'realizado': realizado
                }
            elif agrupar_por_empresa:
                # Agrupar por empresa quando "Todas" empresas foi selecionada
                agrupado_por_empresa_dict[empresa_info][conta][mes] = {
                    'ano_anterior': ano_anterior,
                    'previsto': previsto,
                    'realizado': realizado,
                    'centro_resultado': centro_resultado
                }
                # Acumular totalizadores por empresa
                totalizadores_empresa[empresa_info][f"{mes}_ano_anterior"] += ano_anterior
                totalizadores_empresa[empresa_info][f"{mes}_previsto"] += previsto
                totalizadores_empresa[empresa_info][f"{mes}_realizado"] += realizado
                totalizadores_empresa[empresa_info]["ano_anterior_total"] += ano_anterior
                totalizadores_empresa[empresa_info]["previsto_total"] += previsto
                totalizadores_empresa[empresa_info]["realizado_total"] += realizado
            else:
                # Agrupamento tradicional por conta
                agrupado[conta][mes] = {
                    'ano_anterior': ano_anterior,
                    'previsto': previsto,
                    'realizado': realizado,
                    'centro_resultado': centro_resultado
                }

            # Acumular totalizadores por centro de resultados
            totalizadores_centro[centro_resultado][f"{mes}_ano_anterior"] += ano_anterior
            totalizadores_centro[centro_resultado][f"{mes}_previsto"] += previsto
            totalizadores_centro[centro_resultado][f"{mes}_realizado"] += realizado
            
            # Acumular totais anuais
            totalizadores_centro[centro_resultado]["ano_anterior_total"] += ano_anterior
            totalizadores_centro[centro_resultado]["previsto_total"] += previsto
            totalizadores_centro[centro_resultado]["realizado_total"] += realizado
            
            # Acumular totais anuais por conta (quando não agrupado por centro)
            if not agrupar_por_centro:
                if agrupar_por_empresa:
                    # Quando agrupado por empresa, acumular por empresa+conta
                    totais_anuais_conta[f"{empresa_info}||{conta}"]["ano_anterior_total"] += ano_anterior
                    totais_anuais_conta[f"{empresa_info}||{conta}"]["previsto_total"] += previsto
                    totais_anuais_conta[f"{empresa_info}||{conta}"]["realizado_total"] += realizado
                else:
                    totais_anuais_conta[conta]["ano_anterior_total"] += ano_anterior
                    totais_anuais_conta[conta]["previsto_total"] += previsto
                    totais_anuais_conta[conta]["realizado_total"] += realizado

            # Log para debug da estrutura dos dados (apenas se DEBUG estiver habilitado e para primeiros itens)
            if logger.isEnabledFor(logging.DEBUG) and (len(agrupado) <= 3 if not agrupar_por_centro and not agrupar_por_empresa else len(agrupado_por_centro) <= 3 if agrupar_por_centro else len(agrupado_por_empresa_dict) <= 3):
                logger.debug(f"🔍 Item processado - Empresa: {empresa_info}, Conta: {conta}, Centro: {centro_resultado}, Mês: {mes}")
                logger.debug(f"🔍 Chaves disponíveis no item: {list(item.keys())}")
                logger.debug(f"🔍 Campo centro_resultado encontrado: {centro_resultado}")
                logger.debug(f"🔍 Campo descrcentroresultado: {item.get('descrcentroresultado')}")
                logger.debug(f"🔍 Campo centroresultados: {item.get('centroresultados')}")

        except Exception:
            logger.exception(f"❌ Erro ao agrupar dados para conta {conta} e mês {mes}")

    # Função para converter defaultdict em dict recursivamente
    def to_dict(d):
        if isinstance(d, defaultdict):
            d = {k: to_dict(v) for k, v in d.items()}
        return d

    if agrupar_por_centro:
        logger.info(f"📦 Total de centros agrupados: {len(agrupado_por_centro)}")
        return {
            'agrupado_por_centro': to_dict(agrupado_por_centro),
            'totalizadores_centro': to_dict(totalizadores_centro)
        }
    elif agrupar_por_empresa:
        logger.info(f"📦 Total de empresas agrupadas: {len(agrupado_por_empresa_dict)}")
        
        # Calcular totalizador geral (soma de todas as empresas)
        totalizadores_geral = defaultdict(float)
        for empresa, totais in totalizadores_empresa.items():
            for chave, valor in totais.items():
                totalizadores_geral[chave] += valor
        
        return {
            'agrupado_por_empresa': to_dict(agrupado_por_empresa_dict),
            'totalizadores_centro': to_dict(totalizadores_centro),
            'totais_anuais_conta': to_dict(totais_anuais_conta),
            'totalizadores_empresa': to_dict(totalizadores_empresa),
            'totalizadores_geral': to_dict(totalizadores_geral)
        }
    else:
        logger.info(f"📦 Total de contas agrupadas: {len(agrupado)}")
        return {
            'agrupado': to_dict(agrupado),
            'totalizadores_centro': to_dict(totalizadores_centro),
            'totais_anuais_conta': to_dict(totais_anuais_conta)
        }

# Tela de debug
@token_required
def debug_resultados_view(request):
    resultados = request.session.get('resultados_debug', [])
    return render(request, 'debug_resultados.html', {
        'resultados': resultados
    })

# Tela de configuração
@token_required
def configuracao_view(request):
    token = request.session['token']
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }

    def post_api(endpoint, use_pagination=False):
        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/{endpoint}"
        all_data = []
        
        try:
            if use_pagination:
                # Implementar paginação similar ao painel
                page = 1
                while True:
                    payload = {
                        "page": page,
                        "limit": 1000
                    }
                    resp = requests.post(url, json=payload, headers=headers)
                    
                    if resp.status_code == 200:
                        response_data = resp.json()
                        page_data = response_data.get("data", [])
                        has_next = response_data.get("hasNext", False)
                        
                        all_data.extend(page_data)
                        logger.info(f"📄 Endpoint '{endpoint}' - Página {page}: {len(page_data)} registros")
                        
                        if not has_next:
                            break
                        page += 1
                    else:
                        logger.warning(f"⚠️ Endpoint '{endpoint}' retornou {resp.status_code} na página {page}")
                        break
            else:
                # Sem paginação (para endpoints pequenos)
                resp = requests.post(url, json={"limit": 1000, "page": 1}, headers=headers)
                if resp.status_code == 200:
                    all_data = resp.json().get("data", [])
        except Exception:
            logger.exception(f"❌ Erro ao buscar {endpoint}")
        
        return all_data

    empresas = post_api('cadastro_empresa')
    centros = post_api('cadastro_centroresultados')
    contas = post_api('cadastro_contabil')
    # Não carregar todas as configurações inicialmente - será carregado via AJAX quando filtros forem aplicados
    # configuracoes = post_api('centroresultado_config', use_pagination=True)

    meses = [
        ('1', 'Janeiro'), ('2', 'Fevereiro'), ('3', 'Março'),
        ('4', 'Abril'), ('5', 'Maio'), ('6', 'Junho'),
        ('7', 'Julho'), ('8', 'Agosto'), ('9', 'Setembro'),
        ('10', 'Outubro'), ('11', 'Novembro'), ('12', 'Dezembro'),
    ]

    # Buscar mensagens de sucesso/erro da sessão
    success_message = request.session.pop('success_message', None)
    error_message = request.session.pop('error_message', None)
    
    return render(request, 'configuracao.html', {
        'empresas': empresas,
        'centros': centros,
        'contas': contas,
        'configuracoes': [],  # Inicialmente vazio - será carregado via AJAX
        'meses': meses,
        'success_message': success_message,
        'error_message': error_message
    })

# Salvar configuração
def get_anocadastro(request):
    """Helper para obter ano de cadastro, tratando valores vazios ou '-'"""
    ano = request.POST.get("anocadastro", "").strip()
    if ano and ano != "-" and ano.isdigit():
        return int(ano)
    return None

@token_required
def salvar_configuracao(request):
    if request.method == 'POST':
        headers = {
            'Authorization': f"Bearer {request.session['token']}",
            'Content-Type': 'application/json'
        }

        # Verifica se é uma ação de exclusão
        acao = request.POST.get("acao", "I")  # I = INSERT/UPDATE, D = DELETE
        
        if acao == "D":
            # Exclusão
            dados = {
                "IN_IDEEMPPREVISAO": int(request.POST["empresa"]),
                "IN_IDCTACONTABIL": int(request.POST["conta"]),
                "IN_TIPOPREVISAO": "V",  # Valor padrão para exclusão
                "IN_VALORPREVISAO": 0.0,  # Valor padrão para exclusão
                "IN_IDCENTRORESULTADO": int(request.POST["centro"]),
                "IN_MESPREVISAO": int(request.POST["mes"]) if request.POST.get("mes") else None,
                "IN_ANOCADASTRO": get_anocadastro(request),
                "IN_ACAO": "D"  # Indica ação de DELETE
            }

            logger.info("🗑️ Dados para exclusão:")
            logger.info(f"   - Empresa: {request.POST['empresa']}")
            logger.info(f"   - Centro: {request.POST['centro']}")
            logger.info(f"   - Conta: {request.POST['conta']}")
            logger.info(f"   - Mês: {request.POST.get('mes', 'None')}")
            logger.info(f"   - Ano: {request.POST.get('anocadastro', 'None')}")
            logger.info(f"   - Ação: {acao}")

            logger.info("🗑️ Excluindo configuração:")
            log_json_pretty(dados)

            url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/set_centroresultado_config"
            response = requests.post(url, json=[dados], headers=headers)

            logger.info(f"📥 Status: {response.status_code}")
            logger.info(f"📥 Resposta: {response.text}")

            # Retorna JSON para requisições AJAX (exclusão)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': response.status_code == 200,
                    'message': 'Configuração excluída com sucesso' if response.status_code == 200 else 'Erro ao excluir'
                })
            else:
                return redirect('configuracao')
        else:
            # Inserção/Atualização (lógica original)
            dados = {
                "IN_IDEEMPPREVISAO": int(request.POST["empresa"]),
                "IN_IDCTACONTABIL": int(request.POST["conta"]),
                "IN_TIPOPREVISAO": request.POST["tipo"],
                "IN_VALORPREVISAO": float(request.POST["valor"]),
                "IN_IDCENTRORESULTADO": int(request.POST["centro"]),
                "IN_MESPREVISAO": int(request.POST["mes"]) if request.POST.get("mes") else None,
                "IN_ANOCADASTRO": get_anocadastro(request),
                "IN_ACAO": "I"  # Indica ação de INSERT/UPDATE
            }

            logger.info("📤 Enviando nova configuração:")
            log_json_pretty(dados)

            url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/set_centroresultado_config"
            response = requests.post(url, json=[dados], headers=headers)

            logger.info(f"📥 Status: {response.status_code}")
            logger.info(f"📥 Resposta: {response.text}")

            if response.status_code == 200:
                request.session['success_message'] = 'Configuração salva com sucesso!'
            else:
                request.session['error_message'] = f'Erro ao salvar: {response.text}'

            return redirect('configuracao')

# Atualizar configuração existente (edição inline)
@token_required
def atualizar_configuracao(request):
    if request.method == 'POST':
        headers = {
            'Authorization': f"Bearer {request.session['token']}",
            'Content-Type': 'application/json'
        }

        dados = {
            "IN_IDEEMPPREVISAO": int(request.POST["empresa"]),
            "IN_IDCTACONTABIL": int(request.POST["conta"]),
            "IN_TIPOPREVISAO": request.POST["tipo"],
            "IN_VALORPREVISAO": float(request.POST["valor"]),
            "IN_IDCENTRORESULTADO": int(request.POST["centro"]),
            "IN_MESPREVISAO": int(request.POST["mes"]) if request.POST.get("mes") else None,
            "IN_ANOCADASTRO": int(request.POST["anocadastro"]) if request.POST.get("anocadastro") else None,
            "IN_ACAO": "I"  # Indica ação de INSERT/UPDATE
        }

        logger.info("📤 Atualizando configuração existente:")
        log_json_pretty(dados)

        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/set_centroresultado_config"
        response = requests.post(url, json=[dados], headers=headers)

        logger.info(f"📥 Status: {response.status_code}")
        logger.info(f"📥 Resposta: {response.text}")

        return JsonResponse({
            'success': response.status_code == 200,
            'message': 'Configuração atualizada com sucesso' if response.status_code == 200 else 'Erro ao atualizar'
        })

    return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)

# Buscar configurações via AJAX (filtradas e paginadas)
@token_required
def buscar_configuracoes_ajax(request):
    """View AJAX para buscar configurações com filtros e paginação"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    token = request.session.get('token')
    if not token:
        return JsonResponse({'success': False, 'message': 'Não autenticado'}, status=401)
    
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    
    # Obter filtros da query string
    ano_filter = request.GET.get('ano', '').strip()
    empresa_filter = request.GET.get('empresa', '').strip()
    centro_filter = request.GET.get('centro', '').strip()
    conta_filter = request.GET.get('conta', '').strip()
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    # Construir payload para a API
    clausulas = []
    
    if ano_filter:
        clausulas.append({
            "campo": "anocadastro",
            "operadorlogico": "AND",
            "operador": "IGUAL",
            "valor": int(ano_filter)
        })
    
    if empresa_filter:
        clausulas.append({
            "campo": "ideempprevisao",
            "operadorlogico": "AND",
            "operador": "IGUAL",
            "valor": int(empresa_filter)
        })
    
    if centro_filter:
        clausulas.append({
            "campo": "idcentroresultado",
            "operadorlogico": "AND",
            "operador": "IGUAL",
            "valor": int(centro_filter)
        })
    
    if conta_filter:
        clausulas.append({
            "campo": "idctacontabil",
            "operadorlogico": "AND",
            "operador": "IGUAL",
            "valor": int(conta_filter)
        })
    
    # Se não há filtros, retornar vazio (não carregar tudo)
    if not clausulas:
        return JsonResponse({
            'success': True,
            'data': [],
            'total': 0,
            'page': page,
            'hasNext': False,
            'message': 'Aplique pelo menos um filtro para visualizar as configurações'
        })
    
    payload = {
        "page": page,
        "limit": limit,
        "clausulas": clausulas
    }
    
    url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/centroresultado_config"
    
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        
        if resp.status_code == 200:
            response_data = resp.json()
            configuracoes = response_data.get("data", [])
            total = response_data.get("total", len(configuracoes))
            has_next = response_data.get("hasNext", False)
            
            logger.info(f"📄 Configurações AJAX - Página {page}: {len(configuracoes)} registros (filtros aplicados)")
            
            return JsonResponse({
                'success': True,
                'data': configuracoes,
                'total': total,
                'page': page,
                'hasNext': has_next
            })
        else:
            logger.warning(f"⚠️ Erro ao buscar configurações: {resp.status_code} - {resp.text}")
            return JsonResponse({
                'success': False,
                'message': f'Erro ao buscar configurações: {resp.status_code}'
            }, status=resp.status_code)
    except Exception as e:
        logger.exception(f"❌ Erro ao buscar configurações via AJAX")
        return JsonResponse({
            'success': False,
            'message': f'Erro ao buscar configurações: {str(e)}'
        }, status=500)

# Logout
def logout_view(request):
    request.session.flush()
    return redirect('login')

# Configurar conexão
def configurar_conexao(request):
    if request.method == 'POST':
        try:
            # Parse JSON data
            data = json.loads(request.body)
            api_base_url = data.get('api_base_url')
            
            if not api_base_url:
                return JsonResponse({
                    'success': False, 
                    'message': 'URL da API é obrigatória'
                }, status=400)
            
            # Read current .env file
            env_file_path = '.env'
            try:
                with open(env_file_path, 'r', encoding='utf-8') as f:
                    env_content = f.read()
            except FileNotFoundError:
                # If .env doesn't exist, create from template
                env_content = """# Configurações do Django
DEBUG=False
SECRET_KEY=django-insecure-x1dkg(z5ie%a0!h&dw6ls*ublu7=i@11w)9v6)ithl&+j0@6@-
# Hosts permitidos
ALLOWED_HOSTS=127.0.0.1,localhost
# URL da API externa
API_BASE_URL=http://200.141.41.20:8086
# Credenciais da API
CLIENT_ID=cisspoder-oauth
CLIENT_SECRET=poder7547
API_USERNAME=integracao
API_PASSWORD=13579
# Configurações de segurança
CSRF_TRUSTED_ORIGINS=
# Configurações de logging
LOG_LEVEL=INFO
"""
            
            # Update API_BASE_URL in .env content
            lines = env_content.split('\n')
            updated_lines = []
            api_url_updated = False
            
            for line in lines:
                if line.startswith('API_BASE_URL='):
                    updated_lines.append(f'API_BASE_URL={api_base_url}')
                    api_url_updated = True
                else:
                    updated_lines.append(line)
            
            # If API_BASE_URL wasn't found, add it
            if not api_url_updated:
                # Find the right place to insert (after Django settings)
                insert_index = 0
                for i, line in enumerate(lines):
                    if line.startswith('# URL da API externa') or line.startswith('API_BASE_URL='):
                        insert_index = i + 1
                        break
                
                # Insert the new API_BASE_URL
                if insert_index < len(lines):
                    lines.insert(insert_index, f'API_BASE_URL={api_base_url}')
                else:
                    lines.append(f'API_BASE_URL={api_base_url}')
                
                updated_lines = lines
            
            # Write updated .env file
            with open(env_file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(updated_lines))
            
            logger.info(f"✅ Configuração de conexão atualizada: {api_base_url}")
            
            # Force reload of environment variables
            from decouple import Config, RepositoryEnv
            import os
            
            # Clear any cached config
            if hasattr(config, '_config'):
                delattr(config, '_config')
            
            # Reload environment variables
            try:
                # Force reload by clearing cache
                os.environ.pop('API_BASE_URL', None)
                
                # Reload from .env file
                from decouple import config as reload_config
                reload_config._config = None  # Clear cache
                reload_config._config = RepositoryEnv('.env')
                
                logger.info("🔄 Configurações recarregadas com sucesso!")
                
            except Exception as reload_error:
                logger.warning(f"⚠️ Aviso: Não foi possível recarregar automaticamente. Reinicie a aplicação para aplicar as mudanças. Erro: {reload_error}")
            
            return JsonResponse({
                'success': True,
                'message': 'Configuração salva com sucesso! A aplicação foi recarregada automaticamente.',
                'api_base_url': api_base_url,
                'reloaded': True
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Dados JSON inválidos'
            }, status=400)
        except Exception as e:
            logger.error(f"❌ Erro ao configurar conexão: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Erro interno: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido'
    }, status=405)

# Testar conexão
def testar_conexao(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            api_base_url = data.get('api_base_url')
            
            if not api_base_url:
                return JsonResponse({
                    'success': False,
                    'message': 'URL da API é obrigatória'
                }, status=400)
            
            # Test connection with timeout
            try:
                response = requests.post(
                    f"{api_base_url}/cisspoder-auth/oauth/token",
                    data={
                        'grant_type': 'password',
                        'username': 'test',
                        'password': 'test',
                        'client_id': 'test',
                        'client_secret': 'test'
                    },
                    headers={'Content-Type': 'application/x-www-form-urlencoded'},
                    timeout=5
                )
                
                # If we get any response (even error), connection is working
                if response.status_code in [200, 400, 401, 403]:
                    return JsonResponse({
                        'success': True,
                        'message': 'Conexão estabelecida com sucesso!',
                        'status_code': response.status_code
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'message': f'Conexão estabelecida, mas servidor retornou status {response.status_code}',
                        'status_code': response.status_code
                    })
                    
            except requests.exceptions.Timeout:
                return JsonResponse({
                    'success': False,
                    'message': 'Timeout de conexão - servidor não respondeu em tempo hábil'
                })
            except requests.exceptions.ConnectionError:
                return JsonResponse({
                    'success': False,
                    'message': 'Erro de conexão - verifique se o IP e porta estão corretos'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erro inesperado: {str(e)}'
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Dados JSON inválidos'
            }, status=400)
        except Exception as e:
            logger.error(f"❌ Erro ao testar conexão: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Erro interno: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido'
    }, status=405)

# ==================== GERENCIAMENTO DE USUÁRIOS ====================

# Caminho do arquivo JSON para armazenar configurações de usuários
USUARIOS_JSON_PATH = os.path.join(settings.BASE_DIR, 'usuarios_config.json')

def get_usuarios_config():
    """Lê as configurações de usuários do arquivo JSON"""
    try:
        if os.path.exists(USUARIOS_JSON_PATH):
            with open(USUARIOS_JSON_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"❌ Erro ao ler configurações de usuários: {str(e)}")
    
    # Retorna estrutura padrão se não existir ou houver erro
    return {
        'usuarios_admin': ['ANDERSON.SANTOS'],  # Usuários que podem gerenciar usuários
        'usuarios_config': [],  # Usuários que têm acesso ao botão Configuração
        'usuarios_centros': {}  # Mapeamento usuário -> lista de centros permitidos
    }

def save_usuarios_config(config_data):
    """Salva as configurações de usuários no arquivo JSON"""
    try:
        with open(USUARIOS_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logger.error(f"❌ Erro ao salvar configurações de usuários: {str(e)}")
        return False

def usuario_pode_gerenciar_usuarios(username):
    """Verifica se o usuário pode gerenciar outros usuários"""
    if not username:
        return False
    config_data = get_usuarios_config()
    usuarios_admin = [u.upper() for u in config_data.get('usuarios_admin', [])]
    return username.upper() in usuarios_admin

def usuario_tem_acesso_configuracao(username):
    """Verifica se o usuário tem acesso ao botão de Configuração"""
    if not username:
        return False
    config_data = get_usuarios_config()
    usuarios_config = [u.upper() for u in config_data.get('usuarios_config', [])]
    return username.upper() in usuarios_config


def usuario_pode_aprovar_liberacao(username):
    """Verifica se o usuário pode aprovar/recusar solicitações de liberação de excedente"""
    if not username:
        return False
    config_data = get_usuarios_config()
    aprovadores = [u.upper() for u in config_data.get('usuarios_aprovador_liberacao', [])]
    return username.upper() in aprovadores

def get_centros_permitidos(username):
    """Retorna lista de IDs de centros permitidos para o usuário"""
    if not username:
        return []
    config_data = get_usuarios_config()
    mapping = config_data.get('usuarios_centros', {}) or {}
    centros = mapping.get(username.upper(), [])
    # Garantir ints
    centros_int = []
    for c in centros:
        try:
            centros_int.append(int(c))
        except Exception:
            continue
    return centros_int

# View para tela de configuração de usuários
@token_required
def configuracao_usuarios_view(request):
    """Tela para gerenciar usuários (apenas para admins)"""
    username = request.session.get('username', '')
    
    # Verificar se o usuário tem permissão para acessar esta tela
    if not usuario_pode_gerenciar_usuarios(username):
        return redirect('painel')
    
    # Buscar centros de resultados para permitir vinculação
    token = request.session.get('token')
    centros = []
    if token:
        headers = {
            'Authorization': f"Bearer {token}",
            'Content-Type': 'application/json'
        }
        centros_map = {}
        try:
            url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/cadastro_centroresultados"
            resp = requests.post(url, json={"limit": 1000, "page": 1}, headers=headers, timeout=10)
            if resp.status_code == 200:
                centros = resp.json().get("data", [])
                for c in centros:
                    try:
                        centros_map[int(c.get("idcentroresultado"))] = c.get("centroresultados")
                    except Exception:
                        continue
            else:
                logger.warning(f"⚠️ cadastro_centroresultados retornou {resp.status_code}")
        except Exception:
            logger.exception("❌ Erro ao buscar centros de resultados para configuração de usuários")
    else:
        centros_map = {}
    
    config_data = get_usuarios_config()
    usuarios_admin = config_data.get('usuarios_admin', [])
    usuarios_config = config_data.get('usuarios_config', [])
    usuarios_aprovador_liberacao = config_data.get('usuarios_aprovador_liberacao', [])
    usuarios_centros = config_data.get('usuarios_centros', {})
    usuarios_centros_nomes = {
        user: [centros_map.get(int(cid), str(cid)) for cid in (centros or [])]
        for user, centros in usuarios_centros.items()
    }

    # Buscar mensagens da sessão
    success_message = request.session.pop('success_message', None)
    error_message = request.session.pop('error_message', None)

    return render(request, 'configuracao_usuarios.html', {
        'usuarios_admin': usuarios_admin,
        'usuarios_config': usuarios_config,
        'usuarios_aprovador_liberacao': usuarios_aprovador_liberacao,
        'usuarios_centros': usuarios_centros,
        'usuarios_centros_nomes': usuarios_centros_nomes,
        'centros': centros,
        'centros_map': centros_map,
        'success_message': success_message,
        'error_message': error_message
    })

# View para adicionar/remover usuários admin
@token_required
def gerenciar_usuarios_admin(request):
    """Adiciona ou remove usuários que podem gerenciar outros usuários"""
    username = request.session.get('username', '')
    
    if not usuario_pode_gerenciar_usuarios(username):
        return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)
    
    if request.method == 'POST':
        acao = request.POST.get('acao')  # 'adicionar' ou 'remover'
        usuario = request.POST.get('usuario', '').strip().upper()
        
        if not usuario:
            return JsonResponse({'success': False, 'message': 'Nome do usuário é obrigatório'})
        
        config_data = get_usuarios_config()
        usuarios_admin = config_data.get('usuarios_admin', [])
        usuarios_admin_upper = [u.upper() for u in usuarios_admin]
        
        if acao == 'adicionar':
            if usuario not in usuarios_admin_upper:
                usuarios_admin.append(usuario)
                config_data['usuarios_admin'] = usuarios_admin
                if save_usuarios_config(config_data):
                    request.session['success_message'] = f'Usuário {usuario} adicionado com sucesso!'
                    return JsonResponse({'success': True, 'message': 'Usuário adicionado com sucesso'})
                else:
                    return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})
            else:
                return JsonResponse({'success': False, 'message': 'Usuário já existe na lista'})
        
        elif acao == 'remover':
            if usuario in usuarios_admin_upper:
                usuarios_admin = [u for u in usuarios_admin if u.upper() != usuario]
                config_data['usuarios_admin'] = usuarios_admin
                if save_usuarios_config(config_data):
                    request.session['success_message'] = f'Usuário {usuario} removido com sucesso!'
                    return JsonResponse({'success': True, 'message': 'Usuário removido com sucesso'})
                else:
                    return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})
            else:
                return JsonResponse({'success': False, 'message': 'Usuário não encontrado na lista'})
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)

# View para adicionar/remover usuários com acesso à configuração
@token_required
def gerenciar_usuarios_config(request):
    """Adiciona ou remove usuários que têm acesso ao botão de Configuração"""
    username = request.session.get('username', '')
    
    if not usuario_pode_gerenciar_usuarios(username):
        return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)
    
    if request.method == 'POST':
        acao = request.POST.get('acao')  # 'adicionar' ou 'remover'
        usuario = request.POST.get('usuario', '').strip().upper()
        
        if not usuario:
            return JsonResponse({'success': False, 'message': 'Nome do usuário é obrigatório'})
        
        config_data = get_usuarios_config()
        usuarios_config = config_data.get('usuarios_config', [])
        usuarios_config_upper = [u.upper() for u in usuarios_config]
        
        if acao == 'adicionar':
            if usuario not in usuarios_config_upper:
                usuarios_config.append(usuario)
                config_data['usuarios_config'] = usuarios_config
                if save_usuarios_config(config_data):
                    request.session['success_message'] = f'Usuário {usuario} adicionado com sucesso!'
                    return JsonResponse({'success': True, 'message': 'Usuário adicionado com sucesso'})
                else:
                    return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})
            else:
                return JsonResponse({'success': False, 'message': 'Usuário já existe na lista'})
        
        elif acao == 'remover':
            if usuario in usuarios_config_upper:
                usuarios_config = [u for u in usuarios_config if u.upper() != usuario]
                config_data['usuarios_config'] = usuarios_config
                if save_usuarios_config(config_data):
                    request.session['success_message'] = f'Usuário {usuario} removido com sucesso!'
                    return JsonResponse({'success': True, 'message': 'Usuário removido com sucesso'})
                else:
                    return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})
            else:
                return JsonResponse({'success': False, 'message': 'Usuário não encontrado na lista'})
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)

# View para adicionar/remover aprovadores de liberação de excedente
@token_required
def gerenciar_usuarios_aprovador_liberacao(request):
    """Adiciona ou remove usuários que podem aprovar/recusar solicitações de liberação de excedente"""
    username = request.session.get('username', '')

    if not usuario_pode_gerenciar_usuarios(username):
        return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)

    if request.method == 'POST':
        acao = request.POST.get('acao')  # 'adicionar' ou 'remover'
        usuario = request.POST.get('usuario', '').strip().upper()

        if not usuario:
            return JsonResponse({'success': False, 'message': 'Nome do usuário é obrigatório'})

        config_data = get_usuarios_config()
        lista = config_data.get('usuarios_aprovador_liberacao', [])
        lista_upper = [u.upper() for u in lista]

        if acao == 'adicionar':
            if usuario in lista_upper:
                return JsonResponse({'success': False, 'message': 'Usuário já é aprovador'})
            lista.append(usuario)
            config_data['usuarios_aprovador_liberacao'] = lista
            if save_usuarios_config(config_data):
                return JsonResponse({'success': True, 'message': f'{usuario} agora pode aprovar liberações'})
            return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})

        elif acao == 'remover':
            if usuario not in lista_upper:
                return JsonResponse({'success': False, 'message': 'Usuário não está na lista'})
            lista = [u for u in lista if u.upper() != usuario]
            config_data['usuarios_aprovador_liberacao'] = lista
            if save_usuarios_config(config_data):
                return JsonResponse({'success': True, 'message': f'{usuario} não é mais aprovador'})
            return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})

    return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)


# View para gerenciar vínculo usuário -> centros
@token_required
def gerenciar_usuarios_centros(request):
    """Define lista de centros de resultado permitidos para um usuário"""
    username = request.session.get('username', '')
    
    if not usuario_pode_gerenciar_usuarios(username):
        return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)
    
    if request.method == 'POST':
        usuario = request.POST.get('usuario', '').strip().upper()
        acao = request.POST.get('acao', 'salvar')  # salvar ou limpar
        centros_ids = request.POST.getlist('centros')  # lista de strings
        
        if not usuario:
            return JsonResponse({'success': False, 'message': 'Nome do usuário é obrigatório'})
        
        config_data = get_usuarios_config()
        usuarios_centros = config_data.get('usuarios_centros', {}) or {}
        
        if acao == 'limpar':
            usuarios_centros[usuario] = []
            config_data['usuarios_centros'] = usuarios_centros
            if save_usuarios_config(config_data):
                return JsonResponse({'success': True, 'message': f'Centros limpos para {usuario} (acesso bloqueado).'})
            return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})
        
        # salvar (substitui a lista)
        centros_int = []
        for cid in centros_ids:
            try:
                centros_int.append(int(cid))
            except Exception:
                continue
        
        usuarios_centros[usuario] = centros_int
        config_data['usuarios_centros'] = usuarios_centros
        
        if save_usuarios_config(config_data):
            return JsonResponse({'success': True, 'message': f'Centros atualizados para {usuario}'})
        return JsonResponse({'success': False, 'message': 'Erro ao salvar configuração'})
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)

# Exportar para Excel
@token_required
def exportar_excel_view(request):
    """Exporta os dados do painel para Excel usando dados da sessão (já consultados)"""
    from io import BytesIO
    
    username = request.session.get('username', '')
    centros_permitidos_ids = get_centros_permitidos(username)
    
    # Obter parâmetros da query string (GET) ou POST
    ano = int(request.GET.get('ano', request.POST.get('ano', datetime.now().year)))
    empresa = request.GET.get('empresa') or request.POST.get('empresa') or None
    centro = request.GET.get('centro') or request.POST.get('centro') or None
    
    # Buscar nome do centro e empresa para exibir no Excel
    token = request.session.get('token')
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    
    def post_api(endpoint):
        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/{endpoint}"
        try:
            resp = requests.post(url, json={"page": 1}, headers=headers)
            if resp.status_code == 200:
                return resp.json().get("data", [])
        except Exception:
            logger.exception(f"❌ Erro ao buscar {endpoint}")
        return []
    
    # Buscar lista de centros e empresas para obter nomes
    centros_lista = post_api('cadastro_centroresultados')
    empresas_lista = post_api('cadastro_empresa')
    
    # Filtrar centros conforme permissões
    if centros_permitidos_ids:
        centros_lista = [
            c for c in centros_lista
            if str(c.get('idcentroresultado')) and int(c.get('idcentroresultado')) in centros_permitidos_ids
        ]
    
    # Obter nome do centro selecionado
    nome_centro = None
    if centro and centro.strip():
        try:
            centro_id = int(centro)
            centro_encontrado = next((c for c in centros_lista if c.get('idcentroresultado') == centro_id), None)
            if centro_encontrado:
                nome_centro = centro_encontrado.get('centroresultados') or centro_encontrado.get('descrcentroresultado') or f"Centro {centro_id}"
        except (ValueError, TypeError):
            pass
    
    # Obter nome da empresa selecionada
    nome_empresa = None
    if empresa and empresa.strip():
        try:
            empresa_id = int(empresa)
            empresa_encontrada = next((e for e in empresas_lista if e.get('idempresa') == empresa_id), None)
            if empresa_encontrada:
                nome_empresa = empresa_encontrada.get('empresa') or f"Empresa {empresa_id}"
        except (ValueError, TypeError):
            pass
    
    # Validar permissões de centro
    if centro and centros_permitidos_ids and int(centro) not in centros_permitidos_ids:
        return HttpResponse("Você não tem permissão para acessar este centro de resultado.", status=403)
    
    # Tentar usar dados da sessão (já consultados e processados)
    dados_sessao = request.session.get('dados_exportacao', {})
    
    # Verificar se os dados da sessão correspondem aos filtros atuais
    usar_dados_sessao = (
        dados_sessao and
        dados_sessao.get('ano') == ano and
        str(dados_sessao.get('empresa', '')) == str(empresa or '') and
        str(dados_sessao.get('centro', '')) == str(centro or '')
    )
    
    if usar_dados_sessao:
        # Usar dados já processados da sessão - MUITO MAIS RÁPIDO!
        logger.info("✅ Usando dados da sessão para exportação (sem nova requisição à API)")
        resultado_agrupado = dados_sessao['resultado_agrupado']
        agrupar_por_centro = dados_sessao.get('agrupar_por_centro', False)
        agrupar_por_empresa = dados_sessao.get('agrupar_por_empresa', False)
    else:
        # Fallback: buscar dados da API (caso não tenha dados na sessão)
        logger.info("⚠️ Dados da sessão não encontrados ou filtros diferentes - buscando da API")
        token = request.session['token']
        headers = {
            'Authorization': f"Bearer {token}",
            'Content-Type': 'application/json'
        }
        
        # Construir payload igual ao painel_view
        payload = {
            "page": 1,
            "limit": 1000,
            "clausulas": [
                {"campo": "anoreferencia", "operadorlogico": "AND", "operador": "IGUAL", "valor": ano}
            ]
        }
        
        # Adicionar filtro de empresa
        if empresa and empresa.strip() and empresa != "Todas":
            payload["clausulas"].append({
                "campo": "idempfiltro",
                "operadorlogico": "AND",
                "operador": "IGUAL",
                "valor": int(empresa)
            })
        else:
            payload["clausulas"].append({
                "campo": "idempfiltro",
                "operadorlogico": "AND",
                "operador": "IGUAL",
                "valor": None
            })
        
        # Adicionar filtro de centro
        agrupar_por_centro = False
        agrupar_por_empresa = False
        
        if centro and centro.strip() and centro != "Todos":
            payload["clausulas"].append({
                "campo": "idcentroresultadofiltro",
                "operadorlogico": "AND",
                "operador": "IGUAL",
                "valor": int(centro)
            })
        else:
            if centro == "Todos" or not centro:
                payload["clausulas"].append({
                    "campo": "idcentroresultadofiltro",
                    "operadorlogico": "AND",
                    "operador": "IGUAL",
                    "valor": None
                })
                agrupar_por_centro = True
        
        # Se empresa for "Todas", agrupar por empresa
        if empresa == "Todas" or (not empresa and payload["clausulas"][1]["valor"] is None):
            agrupar_por_empresa = True
            agrupar_por_centro = False
        
        # Buscar dados da API
        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/centro_resultado_bi"
        resultados = []
        
        try:
            page = 1
            while True:
                payload["page"] = page
                resp = requests.post(url, json=payload, headers=headers, timeout=120)
                
                if resp.status_code == 200:
                    data = resp.json()
                    page_data = data.get("data", [])
                    resultados.extend(page_data)
                    
                    if not data.get("hasNext", False):
                        break
                    page += 1
                else:
                    logger.error(f"⚠️ API retornou status {resp.status_code} na página {page}")
                    break
        except requests.exceptions.Timeout:
            logger.exception(f"❌ Timeout ao buscar dados para exportação (mais de 120 segundos)")
            return HttpResponse(
                "<html><body style='font-family: Arial; padding: 20px;'><h1>Erro ao Exportar</h1>"
                "<p>A requisição demorou muito tempo para ser processada. "
                "Tente fazer uma pesquisa primeiro e depois exportar.</p>"
                "<p><a href='javascript:history.back()'>← Voltar</a></p></body></html>",
                status=500
            )
        except Exception as e:
            logger.exception(f"❌ Erro ao buscar dados para exportação: {str(e)}")
            return HttpResponse(
                f"<html><body style='font-family: Arial; padding: 20px;'><h1>Erro ao Exportar</h1>"
                f"<p>Ocorreu um erro ao buscar os dados: {str(e)}</p>"
                "<p><a href='javascript:history.back()'>← Voltar</a></p></body></html>",
                status=500
            )
        
        # Agrupar dados
        resultado_agrupado = agrupar_resultados(resultados, agrupar_por_centro=agrupar_por_centro, agrupar_por_empresa=agrupar_por_empresa)
    
    # Funções auxiliares para determinar cores (mesmas regras da tela)
    def get_previsto_fill(previsto, ano_anterior):
        """Retorna cor laranja claro se previsto > ano_anterior"""
        try:
            previsto_val = float(previsto) if previsto else 0
            ano_ant_val = float(ano_anterior) if ano_anterior else 0
            if previsto_val > ano_ant_val:
                return PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Laranja claro
        except (ValueError, TypeError):
            pass
        return None
    
    def get_realizado_fill(realizado, previsto):
        """Retorna cor verde ou vermelho baseado na comparação"""
        try:
            realizado_val = float(realizado) if realizado else 0
            previsto_val = float(previsto) if previsto else 0
            
            if realizado_val <= 0:
                return None
            
            if realizado_val > previsto_val:
                return PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Vermelho claro (alerta)
            elif realizado_val < previsto_val:
                return PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Verde claro (positivo)
        except (ValueError, TypeError):
            pass
        return None
    
    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle de Orçamento"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    meses_lista = [(f"{i:02d}", meses_nomes[i-1]) for i in range(1, 13)]
    
    row = 1
    
    # Título
    ws.merge_cells(f'A{row}:{get_column_letter(1 + len(meses_lista) * 3 + 2)}{row}')
    cell = ws.cell(row=row, column=1)
    cell.value = f"Controle de Orçamento Empresarial - Ano {ano}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = center_align
    row += 1
    
    # Informações de filtros (Empresa e Centro de Resultado)
    if nome_empresa or nome_centro:
        ws.merge_cells(f'A{row}:{get_column_letter(1 + len(meses_lista) * 3 + 2)}{row}')
        cell = ws.cell(row=row, column=1)
        info_filtros = []
        if nome_empresa:
            info_filtros.append(f"Empresa: {nome_empresa}")
        if nome_centro:
            info_filtros.append(f"Centro de Resultado: {nome_centro}")
        cell.value = " | ".join(info_filtros)
        cell.font = Font(bold=True, size=11, italic=True)
        cell.alignment = center_align
        row += 1
    
    row += 1
    
    # Cabeçalho principal
    col = 1
    ws.cell(row=row, column=col, value="Conta Contábil").fill = header_fill
    ws.cell(row=row, column=col).font = header_font
    ws.cell(row=row, column=col).alignment = center_align
    ws.cell(row=row, column=col).border = border
    col += 1
    
    for mes_num, mes_nome in meses_lista:
        ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+2)}{row}')
        cell = ws.cell(row=row, column=col)
        cell.value = f"{mes_nome} / {ano}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        col += 3
    
    ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+2)}{row}')
    cell = ws.cell(row=row, column=col)
    cell.value = "TOTAIS ANUAIS"
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    row += 1
    
    # Subcabeçalhos
    col = 1
    ws.cell(row=row, column=col, value="").fill = subheader_fill
    ws.cell(row=row, column=col).border = border
    col += 1
    
    for _ in meses_lista:
        ws.cell(row=row, column=col, value="Ano Ant.").fill = subheader_fill
        ws.cell(row=row, column=col).font = subheader_font
        ws.cell(row=row, column=col).alignment = center_align
        ws.cell(row=row, column=col).border = border
        col += 1
        
        ws.cell(row=row, column=col, value="Previsto").fill = subheader_fill
        ws.cell(row=row, column=col).font = subheader_font
        ws.cell(row=row, column=col).alignment = center_align
        ws.cell(row=row, column=col).border = border
        col += 1
        
        ws.cell(row=row, column=col, value="Realizado").fill = subheader_fill
        ws.cell(row=row, column=col).font = subheader_font
        ws.cell(row=row, column=col).alignment = center_align
        ws.cell(row=row, column=col).border = border
        col += 1
    
    for label in ["Ano Ant.", "Previsto", "Realizado"]:
        ws.cell(row=row, column=col, value=label).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        ws.cell(row=row, column=col).font = subheader_font
        ws.cell(row=row, column=col).alignment = center_align
        ws.cell(row=row, column=col).border = border
        col += 1
    
    row += 1
    
    # Dados - Agrupado por empresa
    if agrupar_por_empresa and 'agrupado_por_empresa' in resultado_agrupado:
        agrupado_por_empresa = resultado_agrupado['agrupado_por_empresa']
        totalizadores_empresa = resultado_agrupado.get('totalizadores_empresa', {})
        totalizadores_geral = resultado_agrupado.get('totalizadores_geral', {})
        totais_anuais_conta = resultado_agrupado.get('totais_anuais_conta', {})
        
        for empresa, contas in agrupado_por_empresa.items():
            # Cabeçalho da empresa
            ws.merge_cells(f'A{row}:{get_column_letter(1 + len(meses_lista) * 3 + 2)}{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f"🏢 {empresa}"
            cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            
            # Contas da empresa
            for conta, dados_mes in contas.items():
                col = 1
                ws.cell(row=row, column=col, value=conta).font = Font(bold=True)
                ws.cell(row=row, column=col).border = border
                col += 1
                
                for mes_num, _ in meses_lista:
                    mes_dados = dados_mes.get(mes_num, {})
                    ano_ant = mes_dados.get('ano_anterior', 0) or 0
                    previsto = mes_dados.get('previsto', 0) or 0
                    realizado = mes_dados.get('realizado', 0) or 0
                    
                    # Ano Anterior
                    ws.cell(row=row, column=col, value=float(ano_ant)).number_format = '#,##0.00'
                    ws.cell(row=row, column=col).alignment = right_align
                    ws.cell(row=row, column=col).border = border
                    col += 1
                    
                    # Previsto - aplicar cor laranja se previsto > ano_anterior
                    previsto_cell = ws.cell(row=row, column=col, value=float(previsto))
                    previsto_cell.number_format = '#,##0.00'
                    previsto_cell.alignment = right_align
                    previsto_cell.border = border
                    previsto_fill = get_previsto_fill(previsto, ano_ant)
                    if previsto_fill:
                        previsto_cell.fill = previsto_fill
                    col += 1
                    
                    # Realizado - aplicar cor verde ou vermelho baseado na comparação
                    realizado_cell = ws.cell(row=row, column=col, value=float(realizado))
                    realizado_cell.number_format = '#,##0.00'
                    realizado_cell.alignment = right_align
                    realizado_cell.border = border
                    realizado_fill = get_realizado_fill(realizado, previsto)
                    if realizado_fill:
                        realizado_cell.fill = realizado_fill
                    col += 1
                
                # Totais anuais da conta
                chave_completa = f"{empresa}||{conta}"
                totais_conta = totais_anuais_conta.get(chave_completa, {})
                ano_ant_total = totais_conta.get('ano_anterior_total', 0) or 0
                previsto_total = totais_conta.get('previsto_total', 0) or 0
                realizado_total = totais_conta.get('realizado_total', 0) or 0
                
                # Ano Anterior Total
                ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto Total - aplicar cor laranja se previsto > ano_anterior
                previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
                previsto_total_cell.number_format = '#,##0.00'
                previsto_total_cell.font = Font(bold=True, color="FFFFFF")
                previsto_total_cell.alignment = right_align
                previsto_total_cell.border = border
                # Se previsto > ano_anterior, aplicar laranja sobre azul (mistura)
                if previsto_total > ano_ant_total:
                    previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja mais forte
                else:
                    previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                col += 1
                
                # Realizado Total - aplicar cor verde ou vermelho
                realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
                realizado_total_cell.number_format = '#,##0.00'
                realizado_total_cell.font = Font(bold=True, color="FFFFFF")
                realizado_total_cell.alignment = right_align
                realizado_total_cell.border = border
                # Aplicar cor baseada na comparação
                if realizado_total > previsto_total:
                    realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho mais forte
                elif realizado_total < previsto_total:
                    realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde mais forte
                else:
                    realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                col += 1
                row += 1
            
            # Totalizador da empresa
            ws.merge_cells(f'A{row}:A{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f"💰 TOTAL {empresa}"
            cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.border = border
            col = 2
            
            totais_emp = totalizadores_empresa.get(empresa, {})
            for mes_num, _ in meses_lista:
                chave_ano = f"{mes_num}_ano_anterior"
                chave_prev = f"{mes_num}_previsto"
                chave_real = f"{mes_num}_realizado"
                
                ano_ant_mes = totais_emp.get(chave_ano, 0) or 0
                previsto_mes = totais_emp.get(chave_prev, 0) or 0
                realizado_mes = totais_emp.get(chave_real, 0) or 0
                
                # Ano Anterior Mensal
                ws.cell(row=row, column=col, value=float(ano_ant_mes)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto Mensal - aplicar cor laranja se previsto > ano_anterior
                previsto_mes_cell = ws.cell(row=row, column=col, value=float(previsto_mes))
                previsto_mes_cell.number_format = '#,##0.00'
                previsto_mes_cell.font = Font(bold=True, color="FFFFFF")
                previsto_mes_cell.alignment = right_align
                previsto_mes_cell.border = border
                if previsto_mes > ano_ant_mes:
                    previsto_mes_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Laranja escuro
                else:
                    previsto_mes_cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                col += 1
                
                # Realizado Mensal - aplicar cor verde ou vermelho
                realizado_mes_cell = ws.cell(row=row, column=col, value=float(realizado_mes))
                realizado_mes_cell.number_format = '#,##0.00'
                realizado_mes_cell.font = Font(bold=True, color="FFFFFF")
                realizado_mes_cell.alignment = right_align
                realizado_mes_cell.border = border
                if realizado_mes > previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Vermelho escuro
                elif realizado_mes < previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde escuro
                else:
                    realizado_mes_cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                col += 1
            
            # Totais anuais da empresa
            ano_ant_total = totais_emp.get('ano_anterior_total', 0) or 0
            previsto_total = totais_emp.get('previsto_total', 0) or 0
            realizado_total = totais_emp.get('realizado_total', 0) or 0
            
            # Ano Anterior Total
            ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = right_align
            ws.cell(row=row, column=col).border = border
            col += 1
            
            # Previsto Total - aplicar cor laranja se previsto > ano_anterior
            previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
            previsto_total_cell.number_format = '#,##0.00'
            previsto_total_cell.font = Font(bold=True, color="FFFFFF")
            previsto_total_cell.alignment = right_align
            previsto_total_cell.border = border
            if previsto_total > ano_ant_total:
                previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
            else:
                previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            
            # Realizado Total - aplicar cor verde ou vermelho
            realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
            realizado_total_cell.number_format = '#,##0.00'
            realizado_total_cell.font = Font(bold=True, color="FFFFFF")
            realizado_total_cell.alignment = right_align
            realizado_total_cell.border = border
            if realizado_total > previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho
            elif realizado_total < previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde
            else:
                realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            row += 2
        
        # Totalizador geral
        if totalizadores_geral:
            ws.merge_cells(f'A{row}:A{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = "🌟 TOTAL GERAL (TODAS AS EMPRESAS)"
            cell.fill = PatternFill(start_color="8E44AD", end_color="6C3483", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.border = border
            col = 2
            
            for mes_num, _ in meses_lista:
                chave_ano = f"{mes_num}_ano_anterior"
                chave_prev = f"{mes_num}_previsto"
                chave_real = f"{mes_num}_realizado"
                
                ano_ant_mes = totalizadores_geral.get(chave_ano, 0) or 0
                previsto_mes = totalizadores_geral.get(chave_prev, 0) or 0
                realizado_mes = totalizadores_geral.get(chave_real, 0) or 0
                
                # Ano Anterior Mensal
                ws.cell(row=row, column=col, value=float(ano_ant_mes)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).fill = PatternFill(start_color="8E44AD", end_color="6C3483", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto Mensal - aplicar cor laranja se previsto > ano_anterior
                previsto_mes_cell = ws.cell(row=row, column=col, value=float(previsto_mes))
                previsto_mes_cell.number_format = '#,##0.00'
                previsto_mes_cell.font = Font(bold=True, color="FFFFFF")
                previsto_mes_cell.alignment = right_align
                previsto_mes_cell.border = border
                if previsto_mes > ano_ant_mes:
                    previsto_mes_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Laranja escuro
                else:
                    previsto_mes_cell.fill = PatternFill(start_color="8E44AD", end_color="6C3483", fill_type="solid")
                col += 1
                
                # Realizado Mensal - aplicar cor verde ou vermelho
                realizado_mes_cell = ws.cell(row=row, column=col, value=float(realizado_mes))
                realizado_mes_cell.number_format = '#,##0.00'
                realizado_mes_cell.font = Font(bold=True, color="FFFFFF")
                realizado_mes_cell.alignment = right_align
                realizado_mes_cell.border = border
                if realizado_mes > previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Vermelho escuro
                elif realizado_mes < previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde escuro
                else:
                    realizado_mes_cell.fill = PatternFill(start_color="8E44AD", end_color="6C3483", fill_type="solid")
                col += 1
            
            # Totais anuais gerais
            ano_ant_total = totalizadores_geral.get('ano_anterior_total', 0) or 0
            previsto_total = totalizadores_geral.get('previsto_total', 0) or 0
            realizado_total = totalizadores_geral.get('realizado_total', 0) or 0
            
            # Ano Anterior Total
            ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = right_align
            ws.cell(row=row, column=col).border = border
            col += 1
            
            # Previsto Total - aplicar cor laranja se previsto > ano_anterior
            previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
            previsto_total_cell.number_format = '#,##0.00'
            previsto_total_cell.font = Font(bold=True, color="FFFFFF")
            previsto_total_cell.alignment = right_align
            previsto_total_cell.border = border
            if previsto_total > ano_ant_total:
                previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
            else:
                previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            
            # Realizado Total - aplicar cor verde ou vermelho
            realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
            realizado_total_cell.number_format = '#,##0.00'
            realizado_total_cell.font = Font(bold=True, color="FFFFFF")
            realizado_total_cell.alignment = right_align
            realizado_total_cell.border = border
            if realizado_total > previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho
            elif realizado_total < previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde
            else:
                realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
    
    elif agrupar_por_centro and 'agrupado_por_centro' in resultado_agrupado:
        agrupado_por_centro = resultado_agrupado['agrupado_por_centro']
        totalizadores_centro = resultado_agrupado.get('totalizadores_centro', {})
        totais_anuais_conta = resultado_agrupado.get('totais_anuais_conta', {})
        
        for centro, contas in agrupado_por_centro.items():
            # Cabeçalho do centro
            ws.merge_cells(f'A{row}:{get_column_letter(1 + len(meses_lista) * 3 + 2)}{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f"🏢 {centro}"
            cell.fill = PatternFill(start_color="2C3E50", end_color="34495E", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
            
            # Contas do centro
            for conta, dados_mes in contas.items():
                col = 1
                ws.cell(row=row, column=col, value=conta).font = Font(bold=True)
                ws.cell(row=row, column=col).border = border
                col += 1
                
                for mes_num, _ in meses_lista:
                    mes_dados = dados_mes.get(mes_num, {})
                    ano_ant = mes_dados.get('ano_anterior', 0) or 0
                    previsto = mes_dados.get('previsto', 0) or 0
                    realizado = mes_dados.get('realizado', 0) or 0
                    
                    # Ano Anterior
                    ws.cell(row=row, column=col, value=float(ano_ant)).number_format = '#,##0.00'
                    ws.cell(row=row, column=col).alignment = right_align
                    ws.cell(row=row, column=col).border = border
                    col += 1
                    
                    # Previsto - aplicar cor laranja se previsto > ano_anterior
                    previsto_cell = ws.cell(row=row, column=col, value=float(previsto))
                    previsto_cell.number_format = '#,##0.00'
                    previsto_cell.alignment = right_align
                    previsto_cell.border = border
                    previsto_fill = get_previsto_fill(previsto, ano_ant)
                    if previsto_fill:
                        previsto_cell.fill = previsto_fill
                    col += 1
                    
                    # Realizado - aplicar cor verde ou vermelho baseado na comparação
                    realizado_cell = ws.cell(row=row, column=col, value=float(realizado))
                    realizado_cell.number_format = '#,##0.00'
                    realizado_cell.alignment = right_align
                    realizado_cell.border = border
                    realizado_fill = get_realizado_fill(realizado, previsto)
                    if realizado_fill:
                        realizado_cell.fill = realizado_fill
                    col += 1
                
                # Totais anuais da conta
                totais_conta = totais_anuais_conta.get(conta, {})
                ano_ant_total = totais_conta.get('ano_anterior_total', 0) or 0
                previsto_total = totais_conta.get('previsto_total', 0) or 0
                realizado_total = totais_conta.get('realizado_total', 0) or 0
                
                # Ano Anterior Total
                ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto Total - aplicar cor laranja se previsto > ano_anterior
                previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
                previsto_total_cell.number_format = '#,##0.00'
                previsto_total_cell.font = Font(bold=True, color="FFFFFF")
                previsto_total_cell.alignment = right_align
                previsto_total_cell.border = border
                if previsto_total > ano_ant_total:
                    previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
                else:
                    previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                col += 1
                
                # Realizado Total - aplicar cor verde ou vermelho
                realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
                realizado_total_cell.number_format = '#,##0.00'
                realizado_total_cell.font = Font(bold=True, color="FFFFFF")
                realizado_total_cell.alignment = right_align
                realizado_total_cell.border = border
                if realizado_total > previsto_total:
                    realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho
                elif realizado_total < previsto_total:
                    realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde
                else:
                    realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                col += 1
                row += 1
            
            # Totalizador do centro
            ws.merge_cells(f'A{row}:A{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f"💰 TOTAL {centro}"
            cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.border = border
            col = 2
            
            totais_cent = totalizadores_centro.get(centro, {})
            for mes_num, _ in meses_lista:
                chave_ano = f"{mes_num}_ano_anterior"
                chave_prev = f"{mes_num}_previsto"
                chave_real = f"{mes_num}_realizado"
                
                ano_ant_mes = totais_cent.get(chave_ano, 0) or 0
                previsto_mes = totais_cent.get(chave_prev, 0) or 0
                realizado_mes = totais_cent.get(chave_real, 0) or 0
                
                # Ano Anterior Mensal
                ws.cell(row=row, column=col, value=float(ano_ant_mes)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto Mensal - aplicar cor laranja se previsto > ano_anterior
                previsto_mes_cell = ws.cell(row=row, column=col, value=float(previsto_mes))
                previsto_mes_cell.number_format = '#,##0.00'
                previsto_mes_cell.font = Font(bold=True, color="FFFFFF")
                previsto_mes_cell.alignment = right_align
                previsto_mes_cell.border = border
                if previsto_mes > ano_ant_mes:
                    previsto_mes_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Laranja escuro
                else:
                    previsto_mes_cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                col += 1
                
                # Realizado Mensal - aplicar cor verde ou vermelho
                realizado_mes_cell = ws.cell(row=row, column=col, value=float(realizado_mes))
                realizado_mes_cell.number_format = '#,##0.00'
                realizado_mes_cell.font = Font(bold=True, color="FFFFFF")
                realizado_mes_cell.alignment = right_align
                realizado_mes_cell.border = border
                if realizado_mes > previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Vermelho escuro
                elif realizado_mes < previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde escuro
                else:
                    realizado_mes_cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                col += 1
            
            # Totais anuais do centro
            ano_ant_total = totais_cent.get('ano_anterior_total', 0) or 0
            previsto_total = totais_cent.get('previsto_total', 0) or 0
            realizado_total = totais_cent.get('realizado_total', 0) or 0
            
            # Ano Anterior Total
            ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = right_align
            ws.cell(row=row, column=col).border = border
            col += 1
            
            # Previsto Total - aplicar cor laranja se previsto > ano_anterior
            previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
            previsto_total_cell.number_format = '#,##0.00'
            previsto_total_cell.font = Font(bold=True, color="FFFFFF")
            previsto_total_cell.alignment = right_align
            previsto_total_cell.border = border
            if previsto_total > ano_ant_total:
                previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
            else:
                previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            
            # Realizado Total - aplicar cor verde ou vermelho
            realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
            realizado_total_cell.number_format = '#,##0.00'
            realizado_total_cell.font = Font(bold=True, color="FFFFFF")
            realizado_total_cell.alignment = right_align
            realizado_total_cell.border = border
            if realizado_total > previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho
            elif realizado_total < previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde
            else:
                realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            row += 2
    
    else:
        # Agrupamento simples por conta (uma empresa específica)
        agrupado = resultado_agrupado.get('agrupado', {})
        totais_anuais_conta = resultado_agrupado.get('totais_anuais_conta', {})
        totalizadores_centro = resultado_agrupado.get('totalizadores_centro', {})
        
        for conta, dados_mes in agrupado.items():
            col = 1
            ws.cell(row=row, column=col, value=conta).font = Font(bold=True)
            ws.cell(row=row, column=col).border = border
            col += 1
            
            for mes_num, _ in meses_lista:
                mes_dados = dados_mes.get(mes_num, {})
                ano_ant = mes_dados.get('ano_anterior', 0) or 0
                previsto = mes_dados.get('previsto', 0) or 0
                realizado = mes_dados.get('realizado', 0) or 0
                
                # Ano Anterior
                ws.cell(row=row, column=col, value=float(ano_ant)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto - aplicar cor laranja se previsto > ano_anterior
                previsto_cell = ws.cell(row=row, column=col, value=float(previsto))
                previsto_cell.number_format = '#,##0.00'
                previsto_cell.alignment = right_align
                previsto_cell.border = border
                previsto_fill = get_previsto_fill(previsto, ano_ant)
                if previsto_fill:
                    previsto_cell.fill = previsto_fill
                col += 1
                
                # Realizado - aplicar cor verde ou vermelho baseado na comparação
                realizado_cell = ws.cell(row=row, column=col, value=float(realizado))
                realizado_cell.number_format = '#,##0.00'
                realizado_cell.alignment = right_align
                realizado_cell.border = border
                realizado_fill = get_realizado_fill(realizado, previsto)
                if realizado_fill:
                    realizado_cell.fill = realizado_fill
                col += 1
            
            # Totais anuais da conta
            totais_conta = totais_anuais_conta.get(conta, {})
            ano_ant_total = totais_conta.get('ano_anterior_total', 0) or 0
            previsto_total = totais_conta.get('previsto_total', 0) or 0
            realizado_total = totais_conta.get('realizado_total', 0) or 0
            
            # Ano Anterior Total
            ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = right_align
            ws.cell(row=row, column=col).border = border
            col += 1
            
            # Previsto Total - aplicar cor laranja se previsto > ano_anterior
            previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
            previsto_total_cell.number_format = '#,##0.00'
            previsto_total_cell.font = Font(bold=True, color="FFFFFF")
            previsto_total_cell.alignment = right_align
            previsto_total_cell.border = border
            if previsto_total > ano_ant_total:
                previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
            else:
                previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            
            # Realizado Total - aplicar cor verde ou vermelho
            realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
            realizado_total_cell.number_format = '#,##0.00'
            realizado_total_cell.font = Font(bold=True, color="FFFFFF")
            realizado_total_cell.alignment = right_align
            realizado_total_cell.border = border
            if realizado_total > previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho
            elif realizado_total < previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde
            else:
                realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            row += 1
        
        # Adicionar totalizador quando filtra por uma empresa específica (não agrupado por empresa)
        # Verificar se há totalizadores de centro disponíveis para calcular total geral
        if totalizadores_centro:
            # Calcular totalizador geral somando todos os centros
            totalizador_geral = {}
            for centro, totais in totalizadores_centro.items():
                for chave, valor in totais.items():
                    if chave not in totalizador_geral:
                        totalizador_geral[chave] = 0
                    totalizador_geral[chave] += valor
            
            # Adicionar linha de totalizador
            ws.merge_cells(f'A{row}:A{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = "💰 TOTAL GERAL"
            cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.border = border
            col = 2
            
            for mes_num, _ in meses_lista:
                chave_ano = f"{mes_num}_ano_anterior"
                chave_prev = f"{mes_num}_previsto"
                chave_real = f"{mes_num}_realizado"
                
                ano_ant_mes = totalizador_geral.get(chave_ano, 0) or 0
                previsto_mes = totalizador_geral.get(chave_prev, 0) or 0
                realizado_mes = totalizador_geral.get(chave_real, 0) or 0
                
                # Ano Anterior Mensal
                ws.cell(row=row, column=col, value=float(ano_ant_mes)).number_format = '#,##0.00'
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=col).alignment = right_align
                ws.cell(row=row, column=col).border = border
                col += 1
                
                # Previsto Mensal - aplicar cor laranja se previsto > ano_anterior
                previsto_mes_cell = ws.cell(row=row, column=col, value=float(previsto_mes))
                previsto_mes_cell.number_format = '#,##0.00'
                previsto_mes_cell.font = Font(bold=True, color="FFFFFF")
                previsto_mes_cell.alignment = right_align
                previsto_mes_cell.border = border
                if previsto_mes > ano_ant_mes:
                    previsto_mes_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Laranja escuro
                else:
                    previsto_mes_cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                col += 1
                
                # Realizado Mensal - aplicar cor verde ou vermelho
                realizado_mes_cell = ws.cell(row=row, column=col, value=float(realizado_mes))
                realizado_mes_cell.number_format = '#,##0.00'
                realizado_mes_cell.font = Font(bold=True, color="FFFFFF")
                realizado_mes_cell.alignment = right_align
                realizado_mes_cell.border = border
                if realizado_mes > previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Vermelho escuro
                elif realizado_mes < previsto_mes:
                    realizado_mes_cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde escuro
                else:
                    realizado_mes_cell.fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
                col += 1
            
            # Totais anuais gerais
            ano_ant_total = totalizador_geral.get('ano_anterior_total', 0) or 0
            previsto_total = totalizador_geral.get('previsto_total', 0) or 0
            realizado_total = totalizador_geral.get('realizado_total', 0) or 0
            
            # Ano Anterior Total
            ws.cell(row=row, column=col, value=float(ano_ant_total)).number_format = '#,##0.00'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = right_align
            ws.cell(row=row, column=col).border = border
            col += 1
            
            # Previsto Total - aplicar cor laranja se previsto > ano_anterior
            previsto_total_cell = ws.cell(row=row, column=col, value=float(previsto_total))
            previsto_total_cell.number_format = '#,##0.00'
            previsto_total_cell.font = Font(bold=True, color="FFFFFF")
            previsto_total_cell.alignment = right_align
            previsto_total_cell.border = border
            if previsto_total > ano_ant_total:
                previsto_total_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
            else:
                previsto_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            
            # Realizado Total - aplicar cor verde ou vermelho
            realizado_total_cell = ws.cell(row=row, column=col, value=float(realizado_total))
            realizado_total_cell.number_format = '#,##0.00'
            realizado_total_cell.font = Font(bold=True, color="FFFFFF")
            realizado_total_cell.alignment = right_align
            realizado_total_cell.border = border
            if realizado_total > previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Vermelho
            elif realizado_total < previsto_total:
                realizado_total_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Verde
            else:
                realizado_total_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            col += 1
            row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 40
    for i in range(2, 1 + len(meses_lista) * 3 + 4):
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    # Criar resposta HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="controle_orcamento_{ano}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    return response


# ==================== DRE (Demonstrativo de Resultado do Exercício) ====================

# Cache em memória das listas auxiliares (empresas/centros). Estas listas
# mudam raramente, mas o endpoint da CISS-Poder é lento - cachear evita
# duas chamadas HTTP a cada page load.
_LISTAS_CACHE_TTL_SEG = 3600  # 1 hora
_listas_cache = {"empresas": {"data": None, "ts": 0}, "centros": {"data": None, "ts": 0}}

# Cache dos dados financeiros da DRE por chave ano|empresa|centro (TTL 5 min)
_DRE_DADOS_TTL_SEG = 300
_dre_dados_cache = {}


def _get_lista_cached(nome_cache, endpoint, headers, force_refresh=False):
    """Retorna a lista auxiliar do cache, ou busca da API se expirado."""
    import time
    entry = _listas_cache.get(nome_cache, {"data": None, "ts": 0})
    agora = time.time()
    if (not force_refresh
            and entry["data"] is not None
            and (agora - entry["ts"]) < _LISTAS_CACHE_TTL_SEG):
        logger.info(f"📦 {nome_cache}: cache HIT ({len(entry['data'])} itens, {int(agora - entry['ts'])}s de idade)")
        return entry["data"]

    url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/{endpoint}"
    t0 = time.time()
    try:
        resp = requests.post(url, json={"page": 1}, headers=headers, timeout=60)
        elapsed = time.time() - t0
        if resp.status_code == 200:
            data = resp.json().get("data", [])
            _listas_cache[nome_cache] = {"data": data, "ts": agora}
            logger.info(f"🌐 {nome_cache}: cache MISS - API respondeu em {elapsed:.2f}s ({len(data)} itens)")
            return data
        logger.warning(f"⚠️ {nome_cache}: API retornou {resp.status_code} em {elapsed:.2f}s")
    except Exception:
        logger.exception(f"❌ {nome_cache}: erro ao buscar (após {time.time() - t0:.2f}s)")
    return entry["data"] or []


@token_required
def dre_view(request):
    """Tela DRE: classifica os dados de centro_resultado_bi pela estrutura
    contábil padrão (core/dre_config.py) e exibe linhas + subtotais.

    Por enquanto usa dados mockados. Quando o serviço real for plugado,
    substituir a chamada a gerar_dados_mock pela chamada à API."""
    from datetime import date
    from .dre_service import montar_dre, gerar_dados_mock, _meses_do_periodo
    from .dre_config import get_estrutura_dre

    username = request.session.get('username', '')
    centros_permitidos_ids = get_centros_permitidos(username)
    tem_lista_permitida = bool(centros_permitidos_ids)

    token = request.session['token']
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json',
    }

    empresas = _get_lista_cached('empresas', 'cadastro_empresa', headers)
    centros = _get_lista_cached('centros', 'cadastro_centroresultados', headers)

    if tem_lista_permitida:
        centros = [
            c for c in centros
            if c.get('idcentroresultado') and int(c.get('idcentroresultado')) in centros_permitidos_ids
        ]

    hoje = datetime.now()
    ano_atual = hoje.year
    ano = int(request.GET.get('ano') or request.POST.get('ano') or ano_atual)
    empresa = request.GET.get('empresa') or request.POST.get('empresa') or ''
    centro = request.GET.get('centro') or request.POST.get('centro') or ''
    # Default: dia 1 do mês corrente até o dia de hoje.
    dt_ini_str = request.GET.get('dt_ini') or request.POST.get('dt_ini') or f"{ano}-{hoje.month:02d}-01"
    dt_fim_str = request.GET.get('dt_fim') or request.POST.get('dt_fim') or hoje.strftime("%Y-%m-%d")
    vista = (request.GET.get('vista') or request.POST.get('vista') or 'meses').lower()
    if vista not in ('meses', 'totais'):
        vista = 'meses'

    # Validar permissão de centro
    if centro and tem_lista_permitida:
        try:
            if int(centro) not in centros_permitidos_ids:
                return HttpResponse("Você não tem permissão para acessar este centro de resultado.", status=403)
        except ValueError:
            centro = ''

    # Parse datas
    try:
        dt_ini = date.fromisoformat(dt_ini_str)
    except ValueError:
        dt_ini = date(ano, 1, 1)
    try:
        dt_fim = date.fromisoformat(dt_fim_str)
    except ValueError:
        dt_fim = date(ano, 12, 31)
    if dt_ini > dt_fim:
        dt_ini, dt_fim = dt_fim, dt_ini

    meses = _meses_do_periodo(dt_ini, dt_fim)

    # Só consulta a API quando o usuário clicar em Pesquisar (form submit com pesquisar=1).
    # Primeira visita à tela = só renderiza filtros vazios.
    pesquisar = request.GET.get('pesquisar') or request.POST.get('pesquisar')
    if not pesquisar:
        return render(request, 'dre.html', {
            'empresas': empresas,
            'centros': centros,
            'ano': ano,
            'ano_atual': ano_atual,
            'empresa': empresa,
            'centro': centro,
            'dt_ini': dt_ini.isoformat(),
            'dt_fim': dt_fim.isoformat(),
            'vista': vista,
            'linhas': [],
            'meses': meses,
            'contas_sem_classificacao': [],
            'pode_gerenciar_usuarios': usuario_pode_gerenciar_usuarios(username),
            'tem_acesso_configuracao': usuario_tem_acesso_configuracao(username),
            'total_registros': 0,
            'erro_api': None,
            'consultou': False,
        })

    # Busca os dados no endpoint dedicado da DRE (function UF_CENTRORESULTADOS_DRE).
    # Quando a function nova ainda não estiver publicada na API CISS-Poder,
    # basta trocar DRE_ENDPOINT por 'centro_resultado_bi' no .env para usar a antiga.
    DRE_ENDPOINT = get_dynamic_config('DRE_ENDPOINT', default='centro_resultado_bi_dre')

    # Mês inicial e final extraídos do período selecionado (mesma chave de cache).
    mes_inicial = dt_ini.month
    mes_final = dt_fim.month

    payload = {
        "page": 1,
        "limit": 1000,
        "clausulas": [
            {"campo": "anoreferencia", "operadorlogico": "AND", "operador": "IGUAL", "valor": ano},
            {"campo": "mesinicial",    "operadorlogico": "AND", "operador": "IGUAL", "valor": mes_inicial},
            {"campo": "mesfinal",      "operadorlogico": "AND", "operador": "IGUAL", "valor": mes_final},
        ],
    }
    if empresa and str(empresa).strip():
        try:
            payload["clausulas"].append({
                "campo": "idempfiltro", "operadorlogico": "AND",
                "operador": "IGUAL", "valor": int(empresa),
            })
        except ValueError:
            payload["clausulas"].append({
                "campo": "idempfiltro", "operadorlogico": "AND",
                "operador": "IGUAL", "valor": None,
            })
    else:
        payload["clausulas"].append({
            "campo": "idempfiltro", "operadorlogico": "AND",
            "operador": "IGUAL", "valor": None,
        })

    if centro and str(centro).strip():
        try:
            payload["clausulas"].append({
                "campo": "idcentroresultadofiltro", "operadorlogico": "AND",
                "operador": "IGUAL", "valor": int(centro),
            })
        except ValueError:
            payload["clausulas"].append({
                "campo": "idcentroresultadofiltro", "operadorlogico": "AND",
                "operador": "IGUAL", "valor": None,
            })
    else:
        payload["clausulas"].append({
            "campo": "idcentroresultadofiltro", "operadorlogico": "AND",
            "operador": "IGUAL", "valor": None,
        })

    import time as _time
    # Cache em memória dos dados financeiros - TTL 5 min.
    # Como agora a function aceita filtro de mês, a chave do cache inclui o range.
    cache_key = f"{ano}|{empresa or 'X'}|{centro or 'X'}|{mes_inicial}|{mes_final}"
    cache_entry = _dre_dados_cache.get(cache_key)
    agora_ts = _time.time()
    dados_brutos = []
    erro_api = None
    if cache_entry and (agora_ts - cache_entry["ts"]) < _DRE_DADOS_TTL_SEG:
        dados_brutos = cache_entry["dados"]
        logger.info(f"📦 DRE - cache HIT [{cache_key}] {len(dados_brutos)} registros ({int(agora_ts - cache_entry['ts'])}s de idade)")
    else:
        url = f"{get_dynamic_config('API_BASE_URL')}/cisspoder-service/{DRE_ENDPOINT}"
        t_total = _time.time()
        try:
            page = 1
            while True:
                payload["page"] = page
                t0 = _time.time()
                resp = requests.post(url, json=payload, headers=headers, timeout=120)
                t_page = _time.time() - t0
                if resp.status_code != 200:
                    erro_api = f"API retornou status {resp.status_code}"
                    logger.warning(f"⚠️ DRE - {erro_api} na página {page} ({t_page:.2f}s)")
                    break
                data = resp.json()
                page_data = data.get("data", [])
                dados_brutos.extend(page_data)
                logger.info(f"📄 DRE - página {page}: {len(page_data)} registros em {t_page:.2f}s")
                if not data.get("hasNext", False):
                    break
                page += 1
            logger.info(f"📊 DRE - {len(dados_brutos)} registros totais em {_time.time() - t_total:.2f}s (ano={ano}, empresa={empresa or 'Todas'}, centro={centro or 'Todos'})")
            if not erro_api:
                _dre_dados_cache[cache_key] = {"dados": dados_brutos, "ts": agora_ts}
        except Exception as e:
            erro_api = str(e)
            logger.exception(f"❌ DRE - erro ao buscar centro_resultado_bi (após {_time.time() - t_total:.2f}s)")

    dre = montar_dre(dados_brutos, meses, estrutura=get_estrutura_dre())

    pode_gerenciar_usuarios = usuario_pode_gerenciar_usuarios(username)
    tem_acesso_configuracao = usuario_tem_acesso_configuracao(username)

    context = {
        'empresas': empresas,
        'centros': centros,
        'ano': ano,
        'ano_atual': ano_atual,
        'empresa': empresa,
        'centro': centro,
        'dt_ini': dt_ini.isoformat(),
        'dt_fim': dt_fim.isoformat(),
        'vista': vista,
        'linhas': dre['linhas'],
        'meses': dre['meses'],
        'contas_sem_classificacao': dre['contas_sem_classificacao'],
        'pode_gerenciar_usuarios': pode_gerenciar_usuarios,
        'tem_acesso_configuracao': tem_acesso_configuracao,
        'total_registros': len(dados_brutos),
        'erro_api': erro_api,
        'consultou': True,
    }
    return render(request, 'dre.html', context)
